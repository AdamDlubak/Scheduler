from __future__ import annotations

from datetime import date
from typing import Dict, List, Sequence, Set, Tuple

from openpyxl import Workbook

from scheduler.models import SchedulerConfig, WEEKDAY_NAMES_PL
from scheduler.solver import build_workdays, is_unavailable

EXPECTED_SHEET_HEADERS: Dict[str, Tuple[str, ...]] = {
    "Meta": ("Parametr", "Wartosc"),
    "Pracownicy": ("Pracownik", "Bez_glownego", "Preferuj_mniej_dyzurow"),
    "Swieta": ("Data",),
    "Niedostepnosc_tygodniowa": ("Pracownik", "Dzien_tygodnia"),
    "Wyjatki_dostepnosci": ("Pracownik", "Data"),
    "Urlopy": ("Pracownik", "Data_od", "Data_do"),
    "Wymuszenia": ("Data", "Typ_dyzuru", "Pracownik"),
    "Parametry": ("Parametr", "Wartosc"),
    "Wagi_celu": ("Parametr", "Waga"),
}

REQUIRED_SHEETS = ("Meta", "Pracownicy")
KNOWN_WEIGHT_KEYS = ("total", "main", "extra", "friday", "weekday", "prefer_less_penalty")
KNOWN_PARAM_KEYS = ("solver_max_time_seconds", "solver_num_workers")


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\r", " ").replace("\n", " ").split()).casefold()


def get_sheet_headers(workbook: Workbook, sheet_name: str) -> List[str]:
    if sheet_name not in workbook.sheetnames:
        return []
    worksheet = workbook[sheet_name]
    first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    return ["" if cell is None else str(cell).strip() for cell in first_row]


def validate_sheet_headers(workbook: Workbook, sheet_name: str, errors: List[str]) -> None:
    if sheet_name not in workbook.sheetnames:
        return
    actual_headers = get_sheet_headers(workbook, sheet_name)
    expected_headers = EXPECTED_SHEET_HEADERS[sheet_name]
    normalized_actual = [normalize_header(header) for header in actual_headers]
    normalized_expected = [normalize_header(header) for header in expected_headers]
    if normalized_actual[: len(normalized_expected)] != normalized_expected:
        errors.append(
            f"Arkusz '{sheet_name}' ma niepoprawne naglowki. Oczekiwane pierwsze kolumny: {', '.join(expected_headers)}."
        )


def parse_positive_int(raw: object, field_name: str) -> int:
    if raw is None or str(raw).strip() == "":
        raise ValueError(f"Pole '{field_name}' nie moze byc puste.")
    try:
        value = int(str(raw).strip())
    except ValueError as exc:
        raise ValueError(f"Pole '{field_name}' musi byc liczba calkowita.") from exc
    if value <= 0:
        raise ValueError(f"Pole '{field_name}' musi byc wieksze od zera.")
    return value


def row_has_any_value(row: Sequence[object]) -> bool:
    return any(cell is not None and str(cell).strip() != "" for cell in row)


def day_belongs_to_month(day: date, year: int, month: int) -> bool:
    return day.year == year and day.month == month


def ranges_overlap(a_start: date, a_end: date, b_start: date, b_end: date) -> bool:
    return max(a_start, b_start) <= min(a_end, b_end)


def validate_loaded_config(config: SchedulerConfig) -> tuple[List[str], List[str]]:
    errors: List[str] = []
    warnings: List[str] = []

    try:
        workdays = build_workdays(config.year, config.month, config.holidays)
    except ValueError:
        return errors, warnings

    if not workdays:
        errors.append(f"W miesiacu {config.year}-{config.month:02d} nie ma zadnych dni roboczych po odjeciu swiat.")
        return errors, warnings

    if len(config.employees) < 2:
        errors.append("W arkuszu 'Pracownicy' musza byc co najmniej 2 osoby, bo kazdego dnia sa 2 dyzury.")

    total_slots = len(workdays) * 2
    if len(config.employees) > total_slots:
        warnings.append(
            f"W arkuszu 'Pracownicy' wpisano {len(config.employees)} osob, a w miesiacu {config.year}-{config.month:02d} sa tylko {total_slots} sloty dyzurowe. Czesc osob moze nie dostac zadnego dyzuru."
        )

    for employee in config.employees:
        blocked_weekdays = {wd for wd in config.weekly_unavailability.get(employee, set()) if wd < 5}
        if len(blocked_weekdays) == 5 and not config.weekly_availability_exceptions.get(employee):
            warnings.append(
                f"Pracownik '{employee}' jest niedostepny we wszystkie dni robocze tygodnia wedlug arkusza 'Niedostepnosc_tygodniowa'."
            )

    for (day, shift), employee in config.forced_assignments.items():
        if shift == "glowny" and employee in config.no_main_shift_employees:
            errors.append(f"Wymuszenie: pracownik '{employee}' ma wpisany dyzur glowny {day}, ale ma zakaz dyzurow glownych.")
        if day not in workdays:
            errors.append(f"Wymuszenie {employee} {shift} {day} wypada poza dniem roboczym grafiku.")
            continue
        if is_unavailable(config, employee, day):
            errors.append(f"Wymuszenie {employee} {shift} {day} koliduje z niedostepnoscia lub urlopem.")

    for employee, vacations in config.vacations.items():
        sorted_vacations = sorted(vacations, key=lambda vr: (vr.start, vr.end))
        for idx in range(1, len(sorted_vacations)):
            prev = sorted_vacations[idx - 1]
            curr = sorted_vacations[idx]
            if ranges_overlap(prev.start, prev.end, curr.start, curr.end):
                warnings.append(
                    f"Urlopy pracownika '{employee}' nakladaja sie: {prev.start}–{prev.end} oraz {curr.start}–{curr.end}."
                )
        if sorted_vacations and not any(ranges_overlap(v.start, v.end, workdays[0], workdays[-1]) for v in sorted_vacations):
            warnings.append(f"Pracownik '{employee}' ma urlopy wpisane poza miesiacem {config.year}-{config.month:02d}.")

    for employee, exception_days in config.weekly_availability_exceptions.items():
        for exception_day in sorted(exception_days):
            if any(vr.contains(exception_day) for vr in config.vacations.get(employee, [])):
                warnings.append(
                    f"Pracownik '{employee}' ma wyjatek dostepnosci w dniu {exception_day}, ale tego dnia jest tez na urlopie. Urlop ma pierwszenstwo."
                )

    for day in workdays:
        available = [employee for employee in config.employees if not is_unavailable(config, employee, day)]
        if len(available) < 2:
            errors.append(f"Dzien {day}: dostepne sa tylko {len(available)} osoby, a potrzebne sa 2 dyzury.")

    return errors, warnings


def weekday_name(weekday: int) -> str:
    return WEEKDAY_NAMES_PL.get(weekday, str(weekday))



