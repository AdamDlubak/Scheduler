from __future__ import annotations

from calendar import monthrange
from copy import copy
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Set, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from scheduler.models import (
    ExcelLoadResult,
    InputValidationError,
    SHIFT_EXTRA,
    SHIFT_MAIN,
    ScheduleResult,
    SchedulerConfig,
    VacationRange,
)
from scheduler.validator import (
    KNOWN_PARAM_KEYS,
    KNOWN_WEIGHT_KEYS,
    REQUIRED_SHEETS,
    day_belongs_to_month,
    parse_positive_int,
    row_has_any_value,
    validate_loaded_config,
    validate_sheet_headers,
    weekday_name,
)


POLISH_MONTH_NAMES = {
    1: "STYCZEŃ",
    2: "LUTY",
    3: "MARZEC",
    4: "KWIECIEŃ",
    5: "MAJ",
    6: "CZERWIEC",
    7: "LIPIEC",
    8: "SIERPIEŃ",
    9: "WRZESIEŃ",
    10: "PAŹDZIERNIK",
    11: "LISTOPAD",
    12: "GRUDZIEŃ",
}

POLISH_WEEKDAY_FULL = {
    0: "poniedziałek",
    1: "wtorek",
    2: "środa",
    3: "czwartek",
    4: "piątek",
    5: "sobota",
    6: "niedziela",
}

TEMPLATE_COLUMN_WIDTHS = {
    "A": 8.6640625,
    "B": 12.5,
    "C": 11.6640625,
    "D": 17.6640625,
    "E": 12.5,
    "F": 14.5,
    "G": 21.83203125,
    "H": 26.0,
}

HEADER_FILL = PatternFill(fill_type="solid", start_color="FF808080", end_color="FF808080")
WEEKEND_FILL = PatternFill(fill_type="solid", start_color="FF666666", end_color="FF666666")
HOLIDAY_FILL = PatternFill(fill_type="solid", start_color="FF333333", end_color="FF333333")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER_WRAP_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

CASE_DESCRIPTION_TEXT = "cywilne - rodzinne i opiekuńcze oraz z zakresu postępowań w sprawach nieletnich"
HELP_HEADER_FILL = PatternFill(fill_type="solid", start_color="FFD9EAF7", end_color="FFD9EAF7")
HELP_BODY_FILL = PatternFill(fill_type="solid", start_color="FFF3F8FF", end_color="FFF3F8FF")


def clean_text_value(raw: object) -> str:
    if raw is None:
        return ""
    return " ".join(str(raw).replace("\r", " ").replace("\n", " ").split()).strip()


def row_has_any_value_in_prefix(row: Tuple[object, ...], expected_cols: int) -> bool:
    """Sprawdzaj tylko kolumny danych; kolumny pomocnicze z opisami ignorujemy."""
    return row_has_any_value(row[:expected_cols])


def canonical_text_key(raw: object) -> str:
    return "".join(clean_text_value(raw).split()).casefold()


def resolve_employee_name(
    raw_value: object,
    employee_lookup: Dict[str, str],
    sheet_name: str,
    row_idx: int,
    errors: List[str],
) -> str | None:
    cleaned_value = clean_text_value(raw_value)
    if cleaned_value == "":
        errors.append(f"Arkusz '{sheet_name}', wiersz {row_idx}: kolumna 'Pracownik' nie moze byc pusta.")
        return None
    canonical = canonical_text_key(cleaned_value)
    if canonical not in employee_lookup:
        message = f"Arkusz '{sheet_name}', wiersz {row_idx}: pracownik '{cleaned_value}' nie istnieje w arkuszu 'Pracownicy'."
        if message not in errors:
            errors.append(message)
        return None
    return employee_lookup[canonical]


def validate_employee_references(
    workbook: Workbook,
    employee_lookup: Dict[str, str],
    errors: List[str],
) -> None:
    """Sprawdza, czy pracownicy wpisani w arkuszach pomocniczych istnieją w arkuszu Pracownicy."""
    checks = (
        ("Niedostepnosc_tygodniowa", 0, 2),
        ("Wyjatki_dostepnosci", 0, 2),
        ("Urlopy", 0, 3),
        ("Wymuszenia", 2, 3),
    )
    seen_unknown: Set[Tuple[str, int, str]] = set()

    for sheet_name, employee_col, expected_cols in checks:
        if sheet_name not in workbook.sheetnames:
            continue
        worksheet = workbook[sheet_name]
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row_has_any_value_in_prefix(row, expected_cols):
                continue
            raw_employee = row[employee_col] if len(row) > employee_col else None
            cleaned = clean_text_value(raw_employee)
            if cleaned == "":
                continue
            canonical = canonical_text_key(cleaned)
            if canonical in employee_lookup:
                continue
            key = (sheet_name, row_idx, cleaned)
            if key in seen_unknown:
                continue
            seen_unknown.add(key)
            errors.append(f"Arkusz '{sheet_name}', wiersz {row_idx}: pracownik '{cleaned}' nie istnieje w arkuszu 'Pracownicy'.")


def parse_date_value(raw: object, field_name: str) -> date:
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    if raw is None:
        raise ValueError(f"Brak wartosci daty w polu: {field_name}")

    text = str(raw).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%Y/%m/%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    raise ValueError(f"Niepoprawny format daty '{raw}' w polu: {field_name}")


def parse_bool(raw: object) -> bool:
    if isinstance(raw, bool):
        return raw
    if raw is None:
        return False
    if isinstance(raw, (int, float)):
        return int(raw) != 0
    return clean_text_value(raw).casefold() in {"1", "true", "t", "tak", "yes", "y", "x"}


def parse_weekday(raw: object) -> int:
    if raw is None:
        raise ValueError("Puste pole dnia tygodnia.")
    if isinstance(raw, (int, float)):
        value = int(raw)
        if 1 <= value <= 7:
            return value - 1
        if 0 <= value <= 6:
            return value
        raise ValueError(f"Dzien tygodnia poza zakresem 1-7: {raw}")

    text = clean_text_value(raw).casefold()
    names = {
        "0": 0,
        "1": 0,
        "pon": 0,
        "poniedzialek": 0,
        "2": 1,
        "wt": 1,
        "wtorek": 1,
        "3": 2,
        "sr": 2,
        "sroda": 2,
        "4": 3,
        "czw": 3,
        "czwartek": 3,
        "5": 4,
        "pt": 4,
        "piatek": 4,
        "6": 5,
        "sob": 5,
        "sobota": 5,
        "7": 6,
        "niedz": 6,
        "niedziela": 6,
    }
    if text not in names:
        raise ValueError(f"Nieznany dzien tygodnia: {raw}")
    return names[text]


def parse_shift_value(raw: object) -> str:
    if raw is None:
        raise ValueError("Puste pole typu dyzuru.")
    text = clean_text_value(raw).casefold()
    mapping = {
        SHIFT_MAIN: SHIFT_MAIN,
        "glowny": SHIFT_MAIN,
        "main": SHIFT_MAIN,
        SHIFT_EXTRA: SHIFT_EXTRA,
        "dodatkowy": SHIFT_EXTRA,
        "extra": SHIFT_EXTRA,
    }
    if text not in mapping:
        raise ValueError(f"Nieznany typ dyzuru: {raw}")
    return mapping[text]


def read_kv_sheet(workbook: Workbook, sheet_name: str) -> Dict[str, str]:
    if sheet_name not in workbook.sheetnames:
        return {}
    worksheet = workbook[sheet_name]
    data: Dict[str, str] = {}
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        key = row[0] if len(row) > 0 else None
        value = row[1] if len(row) > 1 else None
        if key is None:
            continue
        data[canonical_text_key(key)] = clean_text_value(value)
    return data


def load_config_from_excel(path: str) -> ExcelLoadResult:
    workbook = load_workbook(path, data_only=True)
    errors: List[str] = []
    warnings: List[str] = []

    for sheet_name in REQUIRED_SHEETS:
        if sheet_name not in workbook.sheetnames:
            errors.append(f"Brak wymaganego arkusza '{sheet_name}'.")
    if errors:
        raise InputValidationError(errors)

    for sheet_name in (
        "Meta",
        "Pracownicy",
        "Swieta",
        "Niedostepnosc_tygodniowa",
        "Wyjatki_dostepnosci",
        "Urlopy",
        "Wymuszenia",
        "Parametry",
        "Wagi_celu",
    ):
        validate_sheet_headers(workbook, sheet_name, errors)
    if errors:
        raise InputValidationError(errors)

    meta = read_kv_sheet(workbook, "Meta")
    try:
        year = parse_positive_int(meta.get("year"), "Meta.year")
    except ValueError as exc:
        errors.append(str(exc))
        year = 1
    try:
        month = parse_positive_int(meta.get("month"), "Meta.month")
    except ValueError as exc:
        errors.append(str(exc))
        month = 1

    if not 1 <= year <= 9999:
        errors.append("Pole 'Meta.year' musi byc liczba od 1 do 9999.")
    if not 1 <= month <= 12:
        errors.append("Pole 'Meta.month' musi byc liczba od 1 do 12.")

    employees: List[str] = []
    no_main_shift_employees: Set[str] = set()
    prefer_less_duties_employees: Set[str] = set()
    seen_employees: Set[str] = set()
    seen_canonical: Dict[str, str] = {}
    employee_lookup: Dict[str, str] = {}

    ws_employees = workbook["Pracownicy"]
    for row_idx, row in enumerate(ws_employees.iter_rows(min_row=2, values_only=True), start=2):
        if not row_has_any_value_in_prefix(row, 3):
            continue
        raw_name = row[0] if len(row) > 0 else None
        cleaned_name = clean_text_value(raw_name)
        if cleaned_name == "":
            errors.append(f"Arkusz 'Pracownicy', wiersz {row_idx}: kolumna 'Pracownik' nie moze byc pusta.")
            continue

        employee = cleaned_name
        if employee in seen_employees:
            errors.append(f"Arkusz 'Pracownicy': pracownik '{employee}' wystepuje wielokrotnie.")
            continue

        normalized = canonical_text_key(employee)
        if normalized in seen_canonical:
            errors.append(
                f"Arkusz 'Pracownicy': nazwy '{seen_canonical[normalized]}' i '{employee}' sa duplikatem (roznia sie tylko wielkoscia liter, spacjami lub enterami)."
            )
            continue
        seen_canonical[normalized] = employee
        employee_lookup[normalized] = employee

        seen_employees.add(employee)
        employees.append(employee)

        no_main = row[1] if len(row) > 1 else None
        prefer_less = row[2] if len(row) > 2 else None
        if parse_bool(no_main):
            no_main_shift_employees.add(employee)
        if parse_bool(prefer_less):
            prefer_less_duties_employees.add(employee)

    if not employees:
        errors.append("Arkusz 'Pracownicy' nie zawiera zadnej aktywnej osoby.")

    if employees:
        validate_employee_references(workbook, employee_lookup, errors)

    holidays: Set[date] = set()
    if "Swieta" in workbook.sheetnames:
        worksheet = workbook["Swieta"]
        seen_holidays: Set[date] = set()
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row_has_any_value_in_prefix(row, 1):
                continue
            raw_day = row[0] if len(row) > 0 else None
            if raw_day is None:
                errors.append(f"Arkusz 'Swieta', wiersz {row_idx}: kolumna 'Data' nie moze byc pusta.")
                continue
            try:
                holiday = parse_date_value(raw_day, f"Swieta.Data (wiersz {row_idx})")
            except ValueError as exc:
                errors.append(str(exc))
                continue

            if not day_belongs_to_month(holiday, year, month):
                errors.append(f"Arkusz 'Swieta', wiersz {row_idx}: data {holiday} nie nalezy do miesiaca {year}-{month:02d}.")
                continue
            if holiday in seen_holidays:
                warnings.append(f"Arkusz 'Swieta': data {holiday} jest wpisana wiecej niz raz.")
            seen_holidays.add(holiday)
            holidays.add(holiday)

    weekly_unavailability: Dict[str, Set[int]] = {employee: set() for employee in employees}
    if "Niedostepnosc_tygodniowa" in workbook.sheetnames:
        worksheet = workbook["Niedostepnosc_tygodniowa"]
        seen_weekly: Set[Tuple[str, int]] = set()
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row_has_any_value_in_prefix(row, 2):
                continue
            raw_employee = row[0] if len(row) > 0 else None
            raw_weekday = row[1] if len(row) > 1 else None
            if raw_employee is None or raw_weekday is None:
                errors.append(f"Arkusz 'Niedostepnosc_tygodniowa', wiersz {row_idx}: wymagane sa kolumny 'Pracownik' i 'Dzien_tygodnia'.")
                continue

            employee = resolve_employee_name(raw_employee, employee_lookup, "Niedostepnosc_tygodniowa", row_idx, errors)
            if employee is None:
                continue

            try:
                weekday = parse_weekday(raw_weekday)
            except ValueError as exc:
                errors.append(f"Arkusz 'Niedostepnosc_tygodniowa', wiersz {row_idx}: {exc}")
                continue

            key = (employee, weekday)
            if key in seen_weekly:
                warnings.append(
                    f"Arkusz 'Niedostepnosc_tygodniowa': duplikat wpisu {employee} / {weekday_name(weekday)}."
                )
            seen_weekly.add(key)
            weekly_unavailability[employee].add(weekday)

    weekly_availability_exceptions: Dict[str, Set[date]] = {}
    if "Wyjatki_dostepnosci" in workbook.sheetnames:
        worksheet = workbook["Wyjatki_dostepnosci"]
        seen_exceptions: Set[Tuple[str, date]] = set()
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row_has_any_value_in_prefix(row, 2):
                continue
            raw_employee = row[0] if len(row) > 0 else None
            raw_day = row[1] if len(row) > 1 else None
            if raw_employee is None or raw_day is None:
                errors.append(f"Arkusz 'Wyjatki_dostepnosci', wiersz {row_idx}: wymagane sa kolumny 'Pracownik' i 'Data'.")
                continue

            employee = resolve_employee_name(raw_employee, employee_lookup, "Wyjatki_dostepnosci", row_idx, errors)
            if employee is None:
                continue
            try:
                day = parse_date_value(raw_day, f"Wyjatki_dostepnosci.Data (wiersz {row_idx})")
            except ValueError as exc:
                errors.append(str(exc))
                continue

            if not day_belongs_to_month(day, year, month):
                errors.append(
                    f"Arkusz 'Wyjatki_dostepnosci', wiersz {row_idx}: data {day} nie nalezy do miesiaca {year}-{month:02d}."
                )
                continue

            key = (employee, day)
            if key in seen_exceptions:
                warnings.append(f"Arkusz 'Wyjatki_dostepnosci': duplikat wpisu {employee} / {day}.")
            seen_exceptions.add(key)
            weekly_availability_exceptions.setdefault(employee, set()).add(day)

    vacations: Dict[str, List[VacationRange]] = {}
    if "Urlopy" in workbook.sheetnames:
        worksheet = workbook["Urlopy"]
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row_has_any_value_in_prefix(row, 3):
                continue
            raw_employee = row[0] if len(row) > 0 else None
            raw_start = row[1] if len(row) > 1 else None
            raw_end = row[2] if len(row) > 2 else None
            if raw_employee is None or raw_start is None or raw_end is None:
                errors.append(f"Arkusz 'Urlopy', wiersz {row_idx}: wymagane sa kolumny 'Pracownik', 'Data_od', 'Data_do'.")
                continue

            employee = resolve_employee_name(raw_employee, employee_lookup, "Urlopy", row_idx, errors)
            if employee is None:
                continue

            try:
                start_day = parse_date_value(raw_start, f"Urlopy.Data_od (wiersz {row_idx})")
                end_day = parse_date_value(raw_end, f"Urlopy.Data_do (wiersz {row_idx})")
            except ValueError as exc:
                errors.append(str(exc))
                continue

            if start_day > end_day:
                errors.append(
                    f"Arkusz 'Urlopy', wiersz {row_idx}: data poczatkowa {start_day} jest pozniejsza niz data koncowa {end_day}."
                )
                continue

            vacations.setdefault(employee, []).append(VacationRange(start=start_day, end=end_day))

    forced_assignments: Dict[Tuple[date, str], str] = {}
    if "Wymuszenia" in workbook.sheetnames:
        worksheet = workbook["Wymuszenia"]
        seen_forced: Dict[Tuple[date, str], str] = {}
        seen_employee_day: Dict[Tuple[str, date], str] = {}
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row_has_any_value_in_prefix(row, 3):
                continue
            raw_day = row[0] if len(row) > 0 else None
            raw_shift = row[1] if len(row) > 1 else None
            raw_employee = row[2] if len(row) > 2 else None
            if raw_day is None or raw_shift is None or raw_employee is None:
                errors.append(f"Arkusz 'Wymuszenia', wiersz {row_idx}: wymagane sa kolumny 'Data', 'Typ_dyzuru', 'Pracownik'.")
                continue

            employee = resolve_employee_name(raw_employee, employee_lookup, "Wymuszenia", row_idx, errors)
            if employee is None:
                continue

            try:
                day = parse_date_value(raw_day, f"Wymuszenia.Data (wiersz {row_idx})")
                shift = parse_shift_value(raw_shift)
            except ValueError as exc:
                errors.append(str(exc))
                continue

            if not day_belongs_to_month(day, year, month):
                errors.append(f"Arkusz 'Wymuszenia', wiersz {row_idx}: data {day} nie nalezy do miesiaca {year}-{month:02d}.")
                continue

            key = (day, shift)
            if key in seen_forced and seen_forced[key] != employee:
                errors.append(
                    f"Arkusz 'Wymuszenia': konflikt dla {day} / {shift}. Wpisano jednoczesnie '{seen_forced[key]}' i '{employee}'."
                )
                continue
            if key in seen_forced and seen_forced[key] == employee:
                warnings.append(f"Arkusz 'Wymuszenia': duplikat wpisu {employee} / {day} / {shift}.")

            employee_day_key = (employee, day)
            if employee_day_key in seen_employee_day and seen_employee_day[employee_day_key] != shift:
                errors.append(f"Arkusz 'Wymuszenia': pracownik '{employee}' ma wpisane dwa dyzury tego samego dnia {day}.")
                continue

            seen_forced[key] = employee
            seen_employee_day[employee_day_key] = shift
            forced_assignments[key] = employee

    solver_max_time_seconds = 240
    solver_num_workers = 8
    if "Parametry" in workbook.sheetnames:
        worksheet = workbook["Parametry"]
        seen_param_keys: Set[str] = set()
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row_has_any_value_in_prefix(row, 2):
                continue
            raw_key = row[0] if len(row) > 0 else None
            raw_value = row[1] if len(row) > 1 else None
            if raw_key is None:
                errors.append(f"Arkusz 'Parametry', wiersz {row_idx}: kolumna 'Parametr' nie moze byc pusta.")
                continue

            key = canonical_text_key(raw_key)
            if key in seen_param_keys:
                warnings.append(f"Arkusz 'Parametry': parametr '{key}' wpisano wiecej niz raz. Obowiazuje ostatnia poprawna wartosc.")
            seen_param_keys.add(key)

            if key not in KNOWN_PARAM_KEYS:
                warnings.append(f"Arkusz 'Parametry': nieznany parametr '{key}' zostanie pominiety.")
                continue

            try:
                parsed_value = parse_positive_int(raw_value, f"Parametry.{key}")
            except ValueError as exc:
                errors.append(str(exc))
                continue

            if key == "solver_max_time_seconds":
                solver_max_time_seconds = parsed_value
            elif key == "solver_num_workers":
                solver_num_workers = parsed_value

    objective_weights = {
        "total": 8,
        "main": 5,
        "extra": 5,
        "friday": 4,
        "weekday": 2,
        "prefer_less_penalty": 20,
    }
    if "Wagi_celu" in workbook.sheetnames:
        worksheet = workbook["Wagi_celu"]
        seen_weight_keys: Set[str] = set()
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row_has_any_value_in_prefix(row, 2):
                continue
            raw_key = row[0] if len(row) > 0 else None
            raw_value = row[1] if len(row) > 1 else None
            if raw_key is None or raw_value is None:
                errors.append(f"Arkusz 'Wagi_celu', wiersz {row_idx}: wymagane sa kolumny 'Parametr' i 'Waga'.")
                continue

            key = canonical_text_key(raw_key)
            if key in seen_weight_keys:
                warnings.append(f"Arkusz 'Wagi_celu': waga '{key}' wpisana wiecej niz raz. Obowiazuje ostatnia poprawna wartosc.")
            seen_weight_keys.add(key)

            if key not in KNOWN_WEIGHT_KEYS:
                warnings.append(f"Arkusz 'Wagi_celu': nieznana waga '{key}' zostanie pominieta.")
                continue

            try:
                objective_weights[key] = parse_positive_int(raw_value, f"Wagi_celu.{key}")
            except ValueError as exc:
                errors.append(str(exc))

    if errors:
        raise InputValidationError(errors, warnings)

    config = SchedulerConfig(
        year=year,
        month=month,
        employees=employees,
        holidays=holidays,
        weekly_unavailability=weekly_unavailability,
        weekly_availability_exceptions=weekly_availability_exceptions,
        vacations=vacations,
        forced_assignments=forced_assignments,
        no_main_shift_employees=no_main_shift_employees,
        prefer_less_duties_employees=prefer_less_duties_employees,
        solver_max_time_seconds=solver_max_time_seconds,
        solver_num_workers=solver_num_workers,
        objective_weights=objective_weights,
    )

    business_errors, business_warnings = validate_loaded_config(config)
    errors.extend(business_errors)
    warnings.extend(business_warnings)
    if errors:
        raise InputValidationError(errors, warnings)

    return ExcelLoadResult(config=config, warnings=warnings)


def add_instruction_sheet(workbook: Workbook) -> None:
    if "Instrukcja" in workbook.sheetnames:
        del workbook["Instrukcja"]
    worksheet = workbook.create_sheet("Instrukcja")
    worksheet.append(["Krok", "Opis"])
    rows = [
        ("1", "W arkuszu Meta ustaw year i month."),
        ("2", "W arkuszu Pracownicy wpisz osoby i opcjonalne flagi (0/1)."),
        ("3", "W arkuszu Urlopy podaj zakresy Data_od i Data_do."),
        ("4", "W arkuszu Niedostepnosc_tygodniowa wpisuj dni jako 1-7 (1=pon, 7=niedz) lub nazwy (pon, wt, sr...)."),
        ("5", "W arkuszu Wymuszenia wpisuj typ dyzuru: glowny albo dodatkowy."),
        ("6", "Uruchom: python3 generator.py --config <plik.xlsx> --validate-only"),
        ("7", "Jesli walidacja jest poprawna, uruchom generowanie grafiku bez --validate-only."),
        ("8", "Przy bledach sprawdz plik diagnostyczny *_diagnostyka.xlsx (kolor czerwony=blad, zolty=ostrzezenie)."),
    ]
    for row in rows:
        worksheet.append(list(row))
    worksheet.column_dimensions["A"].width = 10
    worksheet.column_dimensions["B"].width = 120


def _copy_row_style(worksheet, source_row: int, target_row: int, max_col: int = 8) -> None:
    for col in range(1, max_col + 1):
        source = worksheet.cell(source_row, col)
        target = worksheet.cell(target_row, col)
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)


def _normalize_calendar_block_size(worksheet, first_day_row: int, target_last_day: int) -> None:
    current_last_row = worksheet.max_row
    target_last_row = first_day_row + target_last_day - 1

    if target_last_row > current_last_row:
        rows_to_add = target_last_row - current_last_row
        worksheet.insert_rows(current_last_row + 1, rows_to_add)
        for row in range(current_last_row + 1, target_last_row + 1):
            _copy_row_style(worksheet, current_last_row, row)
    elif target_last_row < current_last_row:
        worksheet.delete_rows(target_last_row + 1, current_last_row - target_last_row)


def _update_description_merge(worksheet, first_day_row: int, target_last_day: int) -> None:
    ranges_to_remove = [
        merged_range
        for merged_range in worksheet.merged_cells.ranges
        if merged_range.min_col == 2 and merged_range.max_col == 2 and merged_range.min_row == first_day_row
    ]
    for merged_range in ranges_to_remove:
        worksheet.unmerge_cells(str(merged_range))

    last_row = first_day_row + target_last_day - 1
    worksheet.merge_cells(f"B{first_day_row}:B{last_row}")


def _fill_template_calendar(
    worksheet,
    result: ScheduleResult,
    config: SchedulerConfig,
    first_day_row: int = 3,
) -> None:
    _, last_day = monthrange(config.year, config.month)
    schedule_map = {day: (main, extra) for day, _weekday_short, main, extra in result.schedule_rows}

    for day_num in range(1, last_day + 1):
        row = first_day_row + day_num - 1
        current_day = date(config.year, config.month, day_num)
        is_planned_day = current_day.weekday() < 5 and current_day not in config.holidays
        main, extra = schedule_map.get(current_day, (None, None)) if is_planned_day else (None, None)

        worksheet.cell(row=row, column=1).value = day_num
        # Kolumna B jest scalona (B3:B..), zapis tylko w pierwszym wierszu.
        worksheet.cell(row=row, column=3).value = 2
        worksheet.cell(row=row, column=4).value = None
        worksheet.cell(row=row, column=5).value = datetime(config.year, config.month, day_num)
        worksheet.cell(row=row, column=6).value = POLISH_WEEKDAY_FULL[current_day.weekday()]
        worksheet.cell(row=row, column=7).value = main
        worksheet.cell(row=row, column=8).value = extra


def _apply_day_colors(worksheet, config: SchedulerConfig, first_day_row: int = 3, max_col: int = 8) -> None:
    _, last_day = monthrange(config.year, config.month)

    for day_num in range(1, last_day + 1):
        row = first_day_row + day_num - 1
        current_day = date(config.year, config.month, day_num)
        if current_day in config.holidays:
            row_fill = HOLIDAY_FILL
        elif current_day.weekday() >= 5:
            row_fill = WEEKEND_FILL
        else:
            row_fill = None

        for col in range(5, max_col + 1):
            cell = worksheet.cell(row=row, column=col)
            if row_fill is None:
                # Dni robocze bez święta pozostają bez tła.
                cell.fill = PatternFill(fill_type=None)
            else:
                cell.fill = copy(row_fill)


def _ensure_wrapping(worksheet, start_row: int, end_row: int, max_col: int = 8) -> None:
    for row in range(start_row, end_row + 1):
        for col in range(1, max_col + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.alignment = copy(CENTER_WRAP_ALIGNMENT)


def _apply_header_style(worksheet, header_row: int = 2, max_col: int = 8) -> None:
    for col in range(1, max_col + 1):
        cell = worksheet.cell(row=header_row, column=col)
        cell.fill = copy(HEADER_FILL)
        cell.alignment = copy(CENTER_WRAP_ALIGNMENT)
        cell.border = copy(THIN_BORDER)


def _apply_grid_borders(worksheet, start_row: int, end_row: int, max_col: int = 8) -> None:
    for row in range(start_row, end_row + 1):
        for col in range(1, max_col + 1):
            worksheet.cell(row=row, column=col).border = copy(THIN_BORDER)


def _set_case_description(worksheet, first_day_row: int = 3) -> None:
    cell = worksheet.cell(row=first_day_row, column=2)
    cell.value = CASE_DESCRIPTION_TEXT
    cell.alignment = copy(CENTER_WRAP_ALIGNMENT)


def _apply_date_format(worksheet, start_row: int, end_row: int, date_col: int = 5) -> None:
    for row in range(start_row, end_row + 1):
        worksheet.cell(row=row, column=date_col).number_format = "[$-415]dd\\ mmmm;@"


def _apply_template_column_widths(worksheet) -> None:
    for col_letter, width in TEMPLATE_COLUMN_WIDTHS.items():
        worksheet.column_dimensions[col_letter].width = width


def _add_sheet_help_box(worksheet, start_col: int, lines: List[Tuple[str, str]]) -> None:
    """Dodaje blok pomocy po prawej stronie arkusza (poza kolumnami parsowanymi)."""
    label_col = start_col
    help_col = start_col + 1
    worksheet.cell(row=1, column=label_col).value = "Pole"
    worksheet.cell(row=1, column=help_col).value = "Opis / format"
    worksheet.cell(row=1, column=label_col).font = Font(bold=True)
    worksheet.cell(row=1, column=help_col).font = Font(bold=True)
    worksheet.cell(row=1, column=label_col).fill = copy(HELP_HEADER_FILL)
    worksheet.cell(row=1, column=help_col).fill = copy(HELP_HEADER_FILL)
    worksheet.cell(row=1, column=label_col).border = copy(THIN_BORDER)
    worksheet.cell(row=1, column=help_col).border = copy(THIN_BORDER)
    worksheet.cell(row=1, column=label_col).alignment = copy(CENTER_WRAP_ALIGNMENT)
    worksheet.cell(row=1, column=help_col).alignment = copy(CENTER_WRAP_ALIGNMENT)

    for idx, (field_name, description) in enumerate(lines, start=2):
        label_cell = worksheet.cell(row=idx, column=label_col)
        help_cell = worksheet.cell(row=idx, column=help_col)
        label_cell.value = field_name
        help_cell.value = description
        label_cell.fill = copy(HELP_BODY_FILL)
        help_cell.fill = copy(HELP_BODY_FILL)
        label_cell.border = copy(THIN_BORDER)
        help_cell.border = copy(THIN_BORDER)
        label_cell.alignment = copy(CENTER_WRAP_ALIGNMENT)
        help_cell.alignment = copy(CENTER_WRAP_ALIGNMENT)

    worksheet.column_dimensions[get_column_letter(label_col)].width = 28
    worksheet.column_dimensions[get_column_letter(help_col)].width = 68


def export_schedule_to_excel(
    result: ScheduleResult,
    output_path: str,
    config: SchedulerConfig,
    template_path: str | None = None,
) -> None:
    default_template = Path(__file__).resolve().parents[1] / "Template.xlsx"
    selected_template = Path(template_path) if template_path else default_template

    if selected_template.exists():
        workbook = load_workbook(selected_template)
        worksheet = workbook[workbook.sheetnames[0]]
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "06.2026"
        worksheet.append([None, None, None, None, None, "ZASTĘPSTWA CZERWIEC 2026 ROK", None, None])
        worksheet.append(
            [
                "l.p.",
                "Rodzaj spraw",
                "Liczba pełniących zastępstwa",
                "Wydział i Sędziowie",
                "DATA",
                "DZIEŃ TYGODNIA",
                "I ZASTĘPCA",
                "II ZASTĘPCA",
            ]
        )
        worksheet.append([1, "cywilne - rodzinne i opiekuncze", 2, None, datetime(2026, 6, 1), "poniedzialek", None, None])

    _, last_day = monthrange(config.year, config.month)

    worksheet.title = f"{config.month:02d}.{config.year}"
    worksheet.cell(row=1, column=6, value=f"ZASTĘPSTWA {POLISH_MONTH_NAMES[config.month]} {config.year} ROK")

    _normalize_calendar_block_size(worksheet, first_day_row=3, target_last_day=last_day)
    _update_description_merge(worksheet, first_day_row=3, target_last_day=last_day)
    _fill_template_calendar(worksheet, result=result, config=config, first_day_row=3)
    _set_case_description(worksheet, first_day_row=3)
    _apply_day_colors(worksheet, config=config, first_day_row=3, max_col=8)
    _apply_template_column_widths(worksheet)
    _ensure_wrapping(worksheet, start_row=2, end_row=2 + last_day, max_col=8)
    _apply_header_style(worksheet, header_row=2, max_col=8)
    _apply_grid_borders(worksheet, start_row=2, end_row=2 + last_day, max_col=8)
    _apply_date_format(worksheet, start_row=3, end_row=2 + last_day, date_col=5)

    workbook.save(output_path)


def export_config_to_excel(config: SchedulerConfig, output_path: str) -> None:
    workbook = Workbook()
    meta = workbook.active
    meta.title = "Meta"
    meta.append(["Parametr", "Wartosc"])
    meta.append(["year", config.year])
    meta.append(["month", config.month])
    _add_sheet_help_box(
        meta,
        start_col=4,
        lines=[
            ("year", "Rok planowania, liczba 4-cyfrowa, np. 2026."),
            ("month", "Miesiac planowania: 1-12 (np. 7 = lipiec)."),
        ],
    )

    employees = workbook.create_sheet("Pracownicy")
    employees.append(["Pracownik", "Bez_glownego", "Preferuj_mniej_dyzurow"])
    for employee in config.employees:
        employees.append(
            [
                employee,
                1 if employee in config.no_main_shift_employees else 0,
                1 if employee in config.prefer_less_duties_employees else 0,
            ]
        )
    _add_sheet_help_box(
        employees,
        start_col=5,
        lines=[
            ("Pracownik", "Imie i nazwisko lub unikalna nazwa osoby."),
            ("Bez_glownego", "0 lub 1. 1 = osoba nie moze pelnic dyzuru glownego."),
            ("Preferuj_mniej_dyzurow", "0 lub 1. 1 = solver stara sie dawac mniej dyzurow tej osobie."),
        ],
    )

    holidays = workbook.create_sheet("Swieta")
    holidays.append(["Data"])
    for holiday in sorted(config.holidays):
        holidays.append([holiday.isoformat()])
    _add_sheet_help_box(
        holidays,
        start_col=3,
        lines=[
            ("Data", "Data swieta do wykluczenia z planowania. Format: RRRR-MM-DD lub DD.MM.RRRR."),
            ("Uwaga", "Podawaj tylko daty z miesiaca ustawionego w arkuszu Meta."),
        ],
    )

    weekly = workbook.create_sheet("Niedostepnosc_tygodniowa")
    weekly.append(["Pracownik", "Dzien_tygodnia"])
    for employee in config.employees:
        for weekday in sorted(config.weekly_unavailability.get(employee, set())):
            weekly.append([employee, weekday + 1])
    _add_sheet_help_box(
        weekly,
        start_col=4,
        lines=[
            ("Pracownik", "Nazwa osoby z arkusza Pracownicy."),
            ("Dzien_tygodnia", "1-7 albo nazwa: pon, wt, sr, czw, pt, sob, niedz."),
            ("Przyklad", "Ania | 1 oznacza niedostepnosc w kazdy poniedzialek."),
        ],
    )

    exceptions = workbook.create_sheet("Wyjatki_dostepnosci")
    exceptions.append(["Pracownik", "Data"])
    for employee, days in sorted(config.weekly_availability_exceptions.items()):
        for day in sorted(days):
            exceptions.append([employee, day.isoformat()])
    _add_sheet_help_box(
        exceptions,
        start_col=4,
        lines=[
            ("Pracownik", "Nazwa osoby z arkusza Pracownicy."),
            ("Data", "Jednorazowy wyjatek dostepnosci. Format daty: RRRR-MM-DD lub DD.MM.RRRR."),
            ("Przyklad", "Osoba ma zwykle wolny poniedzialek, ale 2026-07-06 moze pracowac."),
        ],
    )

    vacations = workbook.create_sheet("Urlopy")
    vacations.append(["Pracownik", "Data_od", "Data_do"])
    for employee, ranges in sorted(config.vacations.items()):
        for vacation in ranges:
            vacations.append([employee, vacation.start.isoformat(), vacation.end.isoformat()])
    _add_sheet_help_box(
        vacations,
        start_col=5,
        lines=[
            ("Pracownik", "Nazwa osoby z arkusza Pracownicy."),
            ("Data_od", "Poczatek urlopu. Format: RRRR-MM-DD lub DD.MM.RRRR."),
            ("Data_do", "Koniec urlopu. Data_do musi byc >= Data_od."),
        ],
    )

    forced = workbook.create_sheet("Wymuszenia")
    forced.append(["Data", "Typ_dyzuru", "Pracownik"])
    for (day, shift), employee in sorted(config.forced_assignments.items(), key=lambda item: (item[0][0], item[0][1])):
        forced.append([day.isoformat(), shift, employee])
    _add_sheet_help_box(
        forced,
        start_col=5,
        lines=[
            ("Data", "Dzien wymuszenia. Format: RRRR-MM-DD lub DD.MM.RRRR."),
            ("Typ_dyzuru", "Dopuszczalne: glowny lub dodatkowy."),
            ("Pracownik", "Nazwa osoby z arkusza Pracownicy."),
        ],
    )

    params = workbook.create_sheet("Parametry")
    params.append(["Parametr", "Wartosc"])
    params.append(["solver_max_time_seconds", config.solver_max_time_seconds])
    params.append(["solver_num_workers", config.solver_num_workers])
    _add_sheet_help_box(
        params,
        start_col=4,
        lines=[
            ("solver_max_time_seconds", "Limit czasu solvera w sekundach. Liczba calkowita > 0."),
            ("solver_num_workers", "Liczba watkow solvera. Liczba calkowita > 0."),
        ],
    )

    weights = workbook.create_sheet("Wagi_celu")
    weights.append(["Parametr", "Waga"])
    for key in KNOWN_WEIGHT_KEYS:
        weights.append([key, int(config.objective_weights.get(key, 0))])
    _add_sheet_help_box(
        weights,
        start_col=4,
        lines=[
            ("Parametr", "Nazwa wagi celu (zostaw domyslne, jesli nie wiesz co zmieniac)."),
            ("Waga", "Liczba calkowita > 0. Wieksza wartosc = wiekszy priorytet kryterium."),
        ],
    )

    add_instruction_sheet(workbook)
    workbook.save(output_path)


def write_template_excel(output_path: str) -> None:
    template = SchedulerConfig(
        year=2026,
        month=6,
        employees=["Pracownik 1", "Pracownik 2"],
        holidays=set(),
        weekly_unavailability={"Pracownik 1": set(), "Pracownik 2": set()},
    )
    export_config_to_excel(template, output_path)


def export_validation_report(
    source_input_path: str,
    report_path: str,
    errors: List[str],
    warnings: List[str],
) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Diagnostyka"
    summary.append(["Poziom", "Komunikat"])

    error_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    warning_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    header_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")

    for cell in summary[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for message in errors:
        summary.append(["BLAD", message])
        row_index = summary.max_row
        summary.cell(row=row_index, column=1).fill = error_fill
        summary.cell(row=row_index, column=2).fill = error_fill

    for message in warnings:
        summary.append(["OSTRZEZENIE", message])
        row_index = summary.max_row
        summary.cell(row=row_index, column=1).fill = warning_fill
        summary.cell(row=row_index, column=2).fill = warning_fill

    summary.column_dimensions["A"].width = 16
    summary.column_dimensions["B"].width = 150

    details = workbook.create_sheet("Szczegoly")
    details.append(["Zrodlo pliku", source_input_path])
    details.append(["Liczba bledow", len(errors)])
    details.append(["Liczba ostrzezen", len(warnings)])

    add_instruction_sheet(workbook)

    Path(report_path).parent.mkdir(parents=True, exist_ok=True)
    workbook.save(report_path)











